# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io


class ExportResultsWizard(models.TransientModel):
    """
    Wizard for exporting batch intake results.
    """
    _name = 'batch.intake.export.wizard'
    _description = 'Export Results Wizard'
    
    batch_id = fields.Many2one(
        'batch.intake',
        string='Batch',
        required=True
    )
    
    export_format = fields.Selection([
        ('xlsx', 'Excel (.xlsx)'),
        ('csv', 'CSV')
    ], string='Export Format', default='xlsx', required=True)
    
    filter_status = fields.Selection([
        ('all', 'All Applicants'),
        ('eligible', 'Eligible Only'),
        ('not_eligible', 'Not Eligible Only'),
        ('pending', 'Pending Only'),
        ('error', 'Errors Only')
    ], string='Filter', default='all', required=True)
    
    include_validation_notes = fields.Boolean(
        string='Include Validation Notes',
        default=True
    )
    
    exported_file = fields.Binary(
        string='Exported File',
        readonly=True
    )
    
    exported_filename = fields.Char(
        string='Filename',
        readonly=True
    )
    
    state = fields.Selection([
        ('draft', 'Draft'),
        ('done', 'Done')
    ], default='draft')
    
    def action_export(self):
        """Export results"""
        self.ensure_one()
        
        # Get applicants based on filter
        domain = [('batch_id', '=', self.batch_id.id)]
        if self.filter_status != 'all':
            domain.append(('eligibility_status', '=', self.filter_status))
        
        applicants = self.env['batch.intake.applicant'].search(domain)
        
        if not applicants:
            raise UserError(_('No applicants found matching the filter criteria.'))
        
        # Generate export file
        if self.export_format == 'xlsx':
            file_data, filename = self._export_excel(applicants)
        else:
            file_data, filename = self._export_csv(applicants)
        
        self.write({
            'exported_file': base64.b64encode(file_data),
            'exported_filename': filename,
            'state': 'done'
        })
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
    
    def _export_excel(self, applicants):
        """Export to Excel"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Batch Results'
        
        # Headers
        headers = ['Name', 'Email', 'Phone', 'Age', 'Nationality', 'Education', 'GPA', 'English Level', 
                   'Eligibility Status', 'Score']
        if self.include_validation_notes:
            headers.append('Validation Notes')
        
        # Style headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row_num, applicant in enumerate(applicants, 2):
            sheet.cell(row=row_num, column=1, value=applicant.name)
            sheet.cell(row=row_num, column=2, value=applicant.email or '')
            sheet.cell(row=row_num, column=3, value=applicant.phone or '')
            sheet.cell(row=row_num, column=4, value=applicant.age or '')
            sheet.cell(row=row_num, column=5, value=applicant.nationality or '')
            sheet.cell(row=row_num, column=6, value=applicant.education_level or '')
            sheet.cell(row=row_num, column=7, value=applicant.gpa or '')
            sheet.cell(row=row_num, column=8, value=applicant.english_level or '')
            sheet.cell(row=row_num, column=9, value=dict(applicant._fields['eligibility_status'].selection).get(applicant.eligibility_status))
            sheet.cell(row=row_num, column=10, value=applicant.eligibility_score)
            
            if self.include_validation_notes:
                sheet.cell(row=row_num, column=11, value=applicant.validation_notes or '')
            
            # Color code status
            status_cell = sheet.cell(row=row_num, column=9)
            if applicant.eligibility_status == 'eligible':
                status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif applicant.eligibility_status == 'not_eligible':
                status_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # Auto-size columns
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        filename = f'batch_{self.batch_id.name}_results.xlsx'
        return output.read(), filename
    
    def _export_csv(self, applicants):
        """Export to CSV"""
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        headers = ['Name', 'Email', 'Phone', 'Age', 'Nationality', 'Education', 'GPA', 'English Level', 
                   'Eligibility Status', 'Score']
        if self.include_validation_notes:
            headers.append('Validation Notes')
        
        writer.writerow(headers)
        
        # Data rows
        for applicant in applicants:
            row = [
                applicant.name,
                applicant.email or '',
                applicant.phone or '',
                applicant.age or '',
                applicant.nationality or '',
                applicant.education_level or '',
                applicant.gpa or '',
                applicant.english_level or '',
                dict(applicant._fields['eligibility_status'].selection).get(applicant.eligibility_status),
                applicant.eligibility_score
            ]
            if self.include_validation_notes:
                row.append(applicant.validation_notes or '')
            
            writer.writerow(row)
        
        filename = f'batch_{self.batch_id.name}_results.csv'
        return output.getvalue().encode('utf-8'), filename

