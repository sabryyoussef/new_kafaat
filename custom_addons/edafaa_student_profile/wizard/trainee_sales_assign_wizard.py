from odoo import _, api, fields, models
from odoo.exceptions import UserError
import base64
import csv
import io
import logging

_logger = logging.getLogger(__name__)


class TraineeSalesAssignWizard(models.TransientModel):
    _name = 'trainee.sales.assign.wizard'
    _description = 'Excel assign trainees to sales staff'

    data_file = fields.Binary(string='Excel file (.xlsx)')
    filename = fields.Char(string='Filename')
    reject_file = fields.Binary(string='Rejected rows', readonly=True)
    reject_filename = fields.Char(readonly=True)
    result_message = fields.Text(string='Result', readonly=True)
    success_count = fields.Integer(readonly=True)
    overwrite_count = fields.Integer(readonly=True)
    reject_count = fields.Integer(readonly=True)

    def action_download_template(self):
        try:
            import openpyxl
        except ImportError as exc:
            raise UserError(_('Python package openpyxl is required: %s') % exc) from exc

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'assignments'
        headers = ['id_number', 'staff_login', 'staff_email', 'trainee_name']
        sheet.append(headers)
        sheet.append(['1099000001', 'admin', 'admin@example.com', 'Sample Trainee'])
        buffer = io.BytesIO()
        workbook.save(buffer)
        data = base64.b64encode(buffer.getvalue())
        attachment = self.env['ir.attachment'].create({
            'name': 'trainee_sales_assign_template.xlsx',
            'type': 'binary',
            'datas': data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id': self.id or 0,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def _require_openpyxl(self):
        try:
            import openpyxl
            return openpyxl
        except ImportError as exc:
            raise UserError(_('Python package openpyxl is required: %s') % exc) from exc

    def _normalize_header(self, value):
        return (str(value or '').strip().lower().replace(' ', '_'))

    def _read_rows(self):
        self.ensure_one()
        if not self.data_file:
            raise UserError(_('Please upload an Excel (.xlsx) file.'))
        name = (self.filename or '').lower()
        if name and not name.endswith('.xlsx'):
            raise UserError(_('Only Excel .xlsx files are supported.'))

        openpyxl = self._require_openpyxl()
        raw = base64.b64decode(self.data_file)
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as exc:
            raise UserError(_('Cannot read Excel file: %s') % exc) from exc
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise UserError(_('The Excel file is empty.')) from exc

        headers = [self._normalize_header(h) for h in header_row]
        required_any = {'id_number'}
        if 'id_number' not in headers:
            raise UserError(_('Missing required column: id_number'))
        if 'staff_login' not in headers and 'staff_email' not in headers:
            raise UserError(_('Missing staff column: provide staff_login and/or staff_email'))

        col_index = {name: idx for idx, name in enumerate(headers) if name}
        data_rows = []
        for row_number, values in enumerate(rows_iter, start=2):
            if not values or all(v is None or str(v).strip() == '' for v in values):
                continue

            def cell(key):
                idx = col_index.get(key)
                if idx is None or idx >= len(values):
                    return ''
                val = values[idx]
                return '' if val is None else str(val).strip()

            data_rows.append({
                'row': row_number,
                'id_number': cell('id_number'),
                'staff_login': cell('staff_login'),
                'staff_email': cell('staff_email'),
                'trainee_name': cell('trainee_name'),
            })
        if not data_rows:
            raise UserError(_('No data rows found in the Excel file.'))
        return data_rows

    def _find_user(self, login, email):
        Users = self.env['res.users'].sudo()
        if login:
            user = Users.search([('login', '=', login)], limit=1)
            if user:
                return user
            user = Users.search([('login', '=ilike', login)], limit=1)
            if user:
                return user
        if email:
            user = Users.search([('email', '=ilike', email)], limit=1)
            if user:
                return user
            user = Users.search([('login', '=ilike', email)], limit=1)
            if user:
                return user
        return Users.browse()

    def action_import(self):
        self.ensure_one()
        rows = self._read_rows()
        Student = self.env['op.student']
        seen_ids = set()
        rejects = []
        success = 0
        overwritten = 0

        for row in rows:
            id_number = row['id_number']
            if not id_number:
                rejects.append((row['row'], id_number, 'missing id_number'))
                continue
            if id_number in seen_ids:
                rejects.append((row['row'], id_number, 'duplicate id_number in file'))
                continue
            seen_ids.add(id_number)

            if not row['staff_login'] and not row['staff_email']:
                rejects.append((row['row'], id_number, 'missing staff_login and staff_email'))
                continue

            student = Student.search([('id_number', '=', id_number)], limit=1)
            if not student:
                rejects.append((row['row'], id_number, 'student not found'))
                continue

            user = self._find_user(row['staff_login'], row['staff_email'])
            if not user:
                rejects.append((row['row'], id_number, 'staff not found'))
                continue

            if student.assigned_user_id and student.assigned_user_id != user:
                overwritten += 1
            student.write({'assigned_user_id': user.id})
            success += 1

        reject_b64 = False
        reject_name = False
        if rejects:
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['row', 'id_number', 'reason'])
            writer.writerows(rejects)
            reject_b64 = base64.b64encode(output.getvalue().encode('utf-8'))
            reject_name = 'trainee_sales_assign_rejects.csv'

        message = _(
            'Import finished: %(success)s assigned (%(overwritten)s overwritten), %(rejected)s rejected.'
        ) % {
            'success': success,
            'overwritten': overwritten,
            'rejected': len(rejects),
        }
        self.write({
            'success_count': success,
            'overwrite_count': overwritten,
            'reject_count': len(rejects),
            'result_message': message,
            'reject_file': reject_b64,
            'reject_filename': reject_name,
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
