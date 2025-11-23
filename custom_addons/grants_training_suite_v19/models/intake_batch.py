# -*- coding: utf-8 -*-

import base64
import csv
import io
import json
import logging
from datetime import datetime
from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)

class IntakeBatch(models.Model):
    _name = 'gr.intake.batch'
    _description = 'Grants Training Intake Batch'
    _order = 'create_date desc'

    # Basic Fields
    name = fields.Char(
        string='Batch Name',
        required=True,
        default=lambda self: _('New'),
        copy=False
    )
    
    filename = fields.Char(
        string='File Name',
        help='Name of the uploaded file'
    )
    
    file_data = fields.Binary(
        string='File',
        help='Upload CSV or Excel file with student data'
    )
    
    file_size = fields.Integer(
        string='File Size (bytes)',
        compute='_compute_file_size',
        store=True,
        help='Size of the uploaded file in bytes'
    )
    
    file_type = fields.Selection([
        ('csv', 'CSV File'),
        ('xlsx', 'Excel File'),
    ], string='File Type', compute='_compute_file_type', store=True)
    
    # Column Mapping Fields
    column_mapping = fields.Text(
        string='Column Mapping',
        help='JSON string containing the mapping between file columns and student fields'
    )
    
    available_columns = fields.Text(
        string='Available Columns',
        help='JSON string containing the detected columns from the uploaded file'
    )
    
    mapping_preview_data = fields.Text(
        string='Mapping Preview Data',
        help='JSON string containing sample data for mapping preview'
    )
    
    # Status Fields
    state = fields.Selection([
        ('draft', 'Draft'),
        ('uploaded', 'File Uploaded'),
        ('mapping', 'Column Mapping'),
        ('validated', 'Validated'),
        ('processed', 'Processed'),
        ('error', 'Error'),
        ('cancelled', 'Cancelled'),
    ], string='Status', default='draft')
    
    # Progress Tracking Fields
    progress_percentage = fields.Float(
        string='Progress (%)',
        compute='_compute_progress_percentage',
        store=True,
        help='Overall progress percentage of the batch processing'
    )
    
    current_stage = fields.Char(
        string='Current Stage',
        compute='_compute_current_stage',
        store=True,
        help='Current processing stage description'
    )
    
    stage_icon = fields.Char(
        string='Stage Icon',
        compute='_compute_stage_icon',
        store=True,
        help='Icon representing the current stage'
    )
    
    # Detailed Progress Tracking
    upload_progress = fields.Selection([
        ('pending', 'Pending'),
        ('in_progress', 'In Progress'),
        ('completed', 'Completed'),
        ('failed', 'Failed'),
    ], string='Upload Progress', default='pending')
    
    mapping_progress = fields.Selection([
        ('pending', 'Pending'),
        ('in_progress', 'In Progress'),
        ('completed', 'Completed'),
        ('failed', 'Failed'),
    ], string='Mapping Progress', default='pending')
    
    validation_progress = fields.Selection([
        ('pending', 'Pending'),
        ('in_progress', 'In Progress'),
        ('completed', 'Completed'),
        ('failed', 'Failed'),
    ], string='Validation Progress', default='pending')
    
    processing_progress = fields.Selection([
        ('pending', 'Pending'),
        ('in_progress', 'In Progress'),
        ('completed', 'Completed'),
        ('failed', 'Failed'),
    ], string='Processing Progress', default='pending')
    
    # Processing Fields
    total_records = fields.Integer(
        string='Total Records',
        help='Total number of records in the file'
    )
    
    processed_records = fields.Integer(
        string='Processed Records',
        default=0,
        help='Number of records successfully processed'
    )
    
    error_records = fields.Integer(
        string='Error Records',
        default=0,
        help='Number of records with errors'
    )
    
    validation_errors = fields.Text(
        string='Validation Errors',
        help='Details of validation errors'
    )
    
    validation_warnings = fields.Text(
        string='Validation Warnings',
        help='Details of validation warnings'
    )
    
    # Import statistics (Phase 2.4)
    created_students_count = fields.Integer(
        string='Students Created',
        default=0,
        help='Number of new students created'
    )
    
    updated_students_count = fields.Integer(
        string='Students Updated',
        default=0,
        help='Number of existing students updated'
    )
    
    import_errors = fields.Text(
        string='Import Errors',
        help='Details of errors during student creation'
    )
    
    import_summary = fields.Text(
        string='Import Summary',
        help='Summary of the import process'
    )
    
    # Failed Records Management (Phase 3.1.2)
    failed_records_data = fields.Text(
        string='Failed Records Data',
        help='JSON string containing failed records with detailed error information'
    )
    
    failed_records_count = fields.Integer(
        string='Failed Records Count',
        default=0,
        compute='_compute_failed_records_count',
        store=True,
        help='Number of records that failed processing'
    )
    
    has_failed_records = fields.Boolean(
        string='Has Failed Records',
        compute='_compute_has_failed_records',
        store=True,
        help='Whether this batch has any failed records'
    )
    
    # Note: Many2one to TransientModel is not allowed in Odoo
    # We'll use a different approach to track correction wizards
    
    # ===== NOTIFICATION FIELDS (Phase 3.1.3) =====
    
    notification_sent = fields.Boolean(
        string='Notification Sent',
        default=False,
        help='Whether notification has been sent for this batch'
    )
    
    notification_type = fields.Selection([
        ('none', 'No Notification'),
        ('success', 'Success Notification'),
        ('error', 'Error Notification'),
        ('warning', 'Warning Notification'),
        ('info', 'Info Notification')
    ], string='Notification Type', default='none', help='Type of notification sent')
    
    notification_message = fields.Text(
        string='Notification Message',
        help='The notification message that was sent'
    )
    
    notification_recipients = fields.Text(
        string='Notification Recipients',
        help='List of users who received notifications'
    )
    
    notification_date = fields.Datetime(
        string='Notification Date',
        help='When the notification was sent'
    )
    
    email_notification_enabled = fields.Boolean(
        string='Email Notifications Enabled',
        default=True,
        help='Whether email notifications are enabled for this batch'
    )
    
    in_app_notification_enabled = fields.Boolean(
        string='In-App Notifications Enabled',
        default=True,
        help='Whether in-app notifications are enabled for this batch'
    )
    
    # ===== SESSION AUTOMATION FIELDS (Phase 3.2) =====
    
    auto_create_sessions = fields.Boolean(
        string='Auto-Create Sessions',
        default=False,
        help='Automatically create course sessions for imported students'
    )
    
    session_creation_enabled = fields.Boolean(
        string='Session Creation Enabled',
        default=True,
        help='Whether session creation is enabled for this batch'
    )
    
    sessions_created_count = fields.Integer(
        string='Sessions Created',
        default=0,
        help='Number of sessions created from this batch'
    )
    
    sessions_scheduled_count = fields.Integer(
        string='Sessions Scheduled',
        default=0,
        help='Number of sessions scheduled from this batch'
    )
    
    session_creation_date = fields.Datetime(
        string='Session Creation Date',
        help='When sessions were created from this batch'
    )
    
    session_creation_errors = fields.Text(
        string='Session Creation Errors',
        help='Any errors encountered during session creation'
    )
    
    session_template_id = fields.Many2one(
        'gr.session.template',
        string='Session Template',
        help='Template to use for creating sessions'
    )
    
    default_session_duration = fields.Float(
        string='Default Session Duration (hours)',
        default=1.0,
        help='Default duration for created sessions'
    )
    
    default_session_type = fields.Selection([
        ('online', 'Online'),
        ('in_person', 'In Person'),
        ('hybrid', 'Hybrid'),
    ], string='Default Session Type', default='online', help='Default type for created sessions')
    
    session_creation_summary = fields.Text(
        string='Session Creation Summary',
        help='Summary of session creation process'
    )
    
    # Timestamps
    upload_date = fields.Datetime(
        string='Upload Date',
        default=fields.Datetime.now
    )
    
    validation_date = fields.Datetime(
        string='Validation Date',
        help='Date when the file was validated'
    )
    
    processing_date = fields.Datetime(
        string='Processing Date',
        help='Date when the file was processed'
    )
    
    # Related Records
    student_ids = fields.One2many(
        'gr.student',
        'intake_batch_id',
        string='Students',
        help='Students created from this batch'
    )
    
    # Computed Fields
    success_rate = fields.Float(
        string='Success Rate (%)',
        compute='_compute_success_rate',
        store=True,
        help='Percentage of successfully processed records'
    )
    
    @api.depends('file_data')
    def _compute_file_size(self):
        """Compute file size from binary data."""
        for record in self:
            if record.file_data:
                # Calculate size from base64 encoded data
                record.file_size = len(record.file_data) * 3 // 4  # Approximate size
            else:
                record.file_size = 0
    
    @api.depends('filename', 'file_data')
    def _compute_file_type(self):
        """Compute file type based on filename extension."""
        for record in self:
            if record.filename:
                filename_lower = record.filename.lower()
                if filename_lower.endswith('.csv'):
                    record.file_type = 'csv'
                elif filename_lower.endswith(('.xlsx', '.xls')):
                    record.file_type = 'xlsx'
                else:
                    record.file_type = False
            else:
                record.file_type = False
    
    @api.depends('total_records', 'processed_records')
    def _compute_success_rate(self):
        """Compute success rate percentage."""
        for record in self:
            if record.total_records > 0:
                record.success_rate = (record.processed_records / record.total_records) * 100
            else:
                record.success_rate = 0.0
    
    @api.depends('state', 'upload_progress', 'mapping_progress', 'validation_progress', 'processing_progress')
    def _compute_progress_percentage(self):
        """Compute overall progress percentage."""
        for record in self:
            progress = 0.0
            
            # Each stage is worth 25% (4 stages total)
            stage_values = {
                'uploaded': 25.0,
                'mapping': 50.0,
                'validated': 75.0,
                'processed': 100.0,
            }
            
            # Base progress from state
            if record.state in stage_values:
                progress = stage_values[record.state]
            elif record.state == 'error':
                # If error, show progress up to the failed stage
                if record.upload_progress == 'failed':
                    progress = 0.0
                elif record.mapping_progress == 'failed':
                    progress = 25.0
                elif record.validation_progress == 'failed':
                    progress = 50.0
                elif record.processing_progress == 'failed':
                    progress = 75.0
                else:
                    progress = 100.0
            elif record.state == 'cancelled':
                progress = 0.0
            
            record.progress_percentage = progress
    
    @api.depends('state', 'upload_progress', 'mapping_progress', 'validation_progress', 'processing_progress')
    def _compute_current_stage(self):
        """Compute current stage description."""
        for record in self:
            if record.state == 'draft':
                record.current_stage = 'Ready to Upload'
            elif record.state == 'uploaded':
                record.current_stage = 'File Uploaded - Ready for Mapping'
            elif record.state == 'mapping':
                record.current_stage = 'Column Mapping Required'
            elif record.state == 'validated':
                record.current_stage = 'File Validated - Ready for Processing'
            elif record.state == 'processed':
                record.current_stage = 'Processing Complete'
            elif record.state == 'error':
                # Determine which stage failed
                if record.upload_progress == 'failed':
                    record.current_stage = 'Upload Failed'
                elif record.mapping_progress == 'failed':
                    record.current_stage = 'Mapping Failed'
                elif record.validation_progress == 'failed':
                    record.current_stage = 'Validation Failed'
                elif record.processing_progress == 'failed':
                    record.current_stage = 'Processing Failed'
                else:
                    record.current_stage = 'Error Occurred'
            elif record.state == 'cancelled':
                record.current_stage = 'Processing Cancelled'
            else:
                record.current_stage = 'Unknown Stage'
    
    @api.depends('state', 'upload_progress', 'mapping_progress', 'validation_progress', 'processing_progress')
    def _compute_stage_icon(self):
        """Compute stage icon."""
        for record in self:
            if record.state == 'draft':
                record.stage_icon = 'fa-upload'
            elif record.state == 'uploaded':
                record.stage_icon = 'fa-file-o'
            elif record.state == 'mapping':
                record.stage_icon = 'fa-map-o'
            elif record.state == 'validated':
                record.stage_icon = 'fa-check-circle'
            elif record.state == 'processed':
                record.stage_icon = 'fa-check'
            elif record.state == 'error':
                record.stage_icon = 'fa-exclamation-triangle'
            elif record.state == 'cancelled':
                record.stage_icon = 'fa-times'
            else:
                record.stage_icon = 'fa-question'
    
    @api.depends('failed_records_data')
    def _compute_failed_records_count(self):
        """Compute the number of failed records."""
        for record in self:
            if record.failed_records_data:
                try:
                    failed_data = json.loads(record.failed_records_data)
                    record.failed_records_count = len(failed_data.get('failed_records', []))
                except (json.JSONDecodeError, TypeError):
                    record.failed_records_count = 0
            else:
                record.failed_records_count = 0
    
    @api.depends('failed_records_count')
    def _compute_has_failed_records(self):
        """Compute whether the batch has any failed records."""
        for record in self:
            record.has_failed_records = record.failed_records_count > 0
    
    
    @api.model
    def create(self, vals_list):
        """Override create to generate sequence number."""
        # Handle both single dict and list of dicts (Odoo 13+)
        if not isinstance(vals_list, list):
            vals_list = [vals_list]
        
        for vals in vals_list:
            if vals.get('name', _('New')) == _('New'):
                vals['name'] = self.env['ir.sequence'].next_by_code('gr.intake.batch') or _('New')
        
        return super(IntakeBatch, self).create(vals_list)
    
    def action_upload_file(self):
        """Action to upload and validate file."""
        self.ensure_one()
        
        if not self.file_data:
            raise UserError(_('Please upload a file first.'))
        
        if not self.filename:
            raise UserError(_('File name is required.'))
        
        # Validate file size (max 10MB)
        max_file_size = 10 * 1024 * 1024  # 10MB in bytes
        if self.file_size > max_file_size:
            raise UserError(_('File size (%d bytes) exceeds maximum allowed size of %d bytes (10MB).') % (self.file_size, max_file_size))
        
        # Validate file type
        if not self.file_type:
            raise UserError(_('Unsupported file type. Please upload CSV (.csv) or Excel (.xlsx, .xls) files only.'))
        
        try:
            # Set upload progress to in_progress
            self.upload_progress = 'in_progress'
            
            # Parse file and count records
            records = self._parse_file()
            self.total_records = len(records)
            
            # Validate that we have records
            if not records:
                self.upload_progress = 'failed'
                raise UserError(_('No records found in the uploaded file. Please check the file format and content.'))
            
            # Update state and progress
            self.state = 'uploaded'
            self.upload_progress = 'completed'
            self.upload_date = fields.Datetime.now()
            
            # Log the action
            _logger.info('File uploaded for batch %s: %s (%d bytes) - %d records found', 
                        self.name, self.filename, self.file_size, self.total_records)
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('File Uploaded Successfully'),
                    'message': _('File "%s" (%d bytes) uploaded successfully. %d records found. You can now validate it.') % (self.filename, self.file_size, self.total_records),
                    'type': 'success',
                }
            }
            
        except Exception as e:
            _logger.error('Error uploading file for batch %s: %s', self.name, str(e))
            self.state = 'error'
            self.upload_progress = 'failed'
            self.validation_errors = str(e)
            raise UserError(_('Error uploading file: %s') % str(e))
    
    def action_validate_file(self):
        """Action to validate the uploaded file (legacy method for direct validation)."""
        self.ensure_one()
        
        if self.state not in ['uploaded', 'mapping']:
            raise UserError(_('Please upload a file first.'))
        
        try:
            # Set validation progress to in_progress
            self.validation_progress = 'in_progress'
            
            # If we have column mapping, use it; otherwise use direct validation
            if self.column_mapping:
                # Use the mapping-based validation
                return self.action_process_with_mapping()
            else:
                # Legacy direct validation (for backward compatibility)
                records = self._parse_file()
                self.total_records = len(records)
                
                # Validate records
                errors = self._validate_records(records)
                
                if errors:
                    self.validation_errors = '\n'.join(errors)
                    self.state = 'error'
                    self.validation_progress = 'failed'
                    self.error_records = len(errors)
                else:
                    self.validation_errors = False
                    self.state = 'validated'
                    self.validation_progress = 'completed'
                    self.error_records = 0
                
                self.validation_date = fields.Datetime.now()
                
                # Log the validation
                _logger.info('File validated for batch %s: %d records, %d errors', 
                            self.name, self.total_records, self.error_records)
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('File Validated'),
                    'message': _('File validation completed. %d records found, %d errors.') % 
                              (self.total_records, self.error_records),
                    'type': 'success' if not errors else 'warning',
                }
            }
            
        except Exception as e:
            _logger.error('Error validating batch %s: %s', self.name, str(e))
            self.state = 'error'
            self.validation_progress = 'failed'
            self.validation_errors = str(e)
            raise UserError(_('Error validating file: %s') % str(e))
    
    def action_process_file(self):
        """Action to process the validated file and create students."""
        self.ensure_one()
        
        _logger.info('Starting file processing for batch %s (state: %s)', self.name, self.state)
        
        if self.state != 'validated':
            raise UserError(_('Please validate the file first.'))
        
        try:
            # Set processing progress to in_progress
            self.processing_progress = 'in_progress'
            
            # Parse file again
            _logger.info('Parsing file for batch %s', self.name)
            records = self._parse_file()
            _logger.info('File parsed successfully: %d records found', len(records))
            
            # Create students
            _logger.info('Creating students for batch %s', self.name)
            created_students = self._create_students(records)
            
            # Update counters and progress
            self.processed_records = len(created_students) + self.updated_students_count
            self.processing_date = fields.Datetime.now()
            self.state = 'processed'
            self.processing_progress = 'completed'
            
            # Log the processing
            _logger.info('File processed for batch %s: %d created, %d updated', 
                        self.name, self.created_students_count, self.updated_students_count)
            
            # Send success notification
            error_count = len(self.import_errors.split(chr(10))) if self.import_errors else 0
            notification_type = 'warning' if error_count > 0 else 'success'
            
            success_message = f"Batch '{self.name}' has been successfully processed. " \
                            f"{self.created_students_count} students created, {self.updated_students_count} updated."
            if error_count > 0:
                success_message += f" {error_count} errors encountered."
            
            success_details = {
                'total_records': len(records),
                'students_created': self.created_students_count,
                'students_updated': self.updated_students_count,
                'errors': error_count,
                'processing_time': str(fields.Datetime.now() - self.upload_date) if self.upload_date else 'Unknown'
            }
            
            self._send_batch_notification(notification_type, success_message, success_details)
            
            # Prepare success message with statistics
            message = f"Import completed successfully!\n"
            message += f"✅ {self.created_students_count} students created\n"
            message += f"🔄 {self.updated_students_count} students updated\n"
            if self.import_errors:
                message += f"❌ {len(self.import_errors.split(chr(10)))} errors encountered\n"
            message += f"📊 Total processed: {self.processed_records} records"
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Import Completed'),
                    'message': message,
                    'type': 'success' if not self.import_errors else 'warning',
                    'sticky': True,
                }
            }
            
        except Exception as e:
            _logger.error('Error processing batch %s: %s', self.name, str(e))
            self.state = 'error'
            self.processing_progress = 'failed'
            self.validation_errors = str(e)
            
            # Send error notification
            error_message = f"Batch '{self.name}' processing failed with error: {str(e)}"
            error_details = {
                'error_type': 'Processing Error',
                'error_message': str(e),
                'batch_status': self.state,
                'failed_at': fields.Datetime.now().isoformat(),
                'total_records': getattr(self, 'total_records', 0),
            }
            
            self._send_batch_notification('error', error_message, error_details)
            
            raise UserError(_('Error processing file: %s') % str(e))
    
    def _parse_file(self):
        """Parse the uploaded file and return records."""
        if not self.file_data:
            return []
        
        try:
            # Decode file data
            file_data = base64.b64decode(self.file_data)
            
            _logger.info('Parsing file: filename=%s, file_type=%s', self.filename, self.file_type)
            
            if self.file_type == 'csv':
                return self._parse_csv(file_data)
            elif self.file_type == 'xlsx':
                return self._parse_excel(file_data)
            else:
                _logger.error('Unsupported file type: %s (filename: %s)', self.file_type, self.filename)
                raise UserError(_('Unsupported file type: %s') % self.file_type)
                
        except Exception as e:
            _logger.error('Error parsing file: %s', str(e))
            raise UserError(_('Error parsing file: %s') % str(e))
    
    def _parse_csv(self, file_data):
        """Parse CSV file and return records."""
        try:
            # Convert bytes to string
            file_content = file_data.decode('utf-8')
            _logger.info('CSV content length: %d characters', len(file_content))
            
            # Parse CSV
            csv_reader = csv.DictReader(io.StringIO(file_content))
            records = list(csv_reader)
            
            _logger.info('CSV parsed successfully: %d records found', len(records))
            if records:
                _logger.info('CSV headers: %s', list(records[0].keys()))
                # Debug: Log the first record to see what data is being parsed
                _logger.info('First record data: %s', records[0])
            
            return records
            
        except Exception as e:
            _logger.error('Error parsing CSV: %s', str(e))
            raise UserError(_('Error parsing CSV file: %s') % str(e))
    
    def _parse_excel(self, file_data):
        """Parse Excel file and return records."""
        try:
            # Determine file format and choose appropriate parser
            if self.filename.endswith('.xlsx'):
                # For .xlsx files, try pandas with openpyxl first
                try:
                    import pandas as pd
                    _logger.info('Parsing .xlsx file using pandas with openpyxl')
                    
                    excel_file = io.BytesIO(file_data)
                    df = pd.read_excel(excel_file, engine='openpyxl')
                    records = df.fillna('').to_dict('records')
                    
                    _logger.info('Excel parsed successfully: %d records found', len(records))
                    if records:
                        _logger.info('Excel headers: %s', list(records[0].keys()))
                    
                    return records
                    
                except ImportError:
                    raise UserError(_('Excel .xlsx files require pandas and openpyxl libraries. Please install: pip install pandas openpyxl'))
                except Exception as e:
                    raise UserError(_('Error parsing .xlsx file with pandas: %s') % str(e))
                    
            elif self.filename.endswith('.xls'):
                # For .xls files, try pandas with xlrd first, then fallback to xlrd only
                try:
                    import pandas as pd
                    _logger.info('Parsing .xls file using pandas with xlrd')
                    
                    excel_file = io.BytesIO(file_data)
                    df = pd.read_excel(excel_file, engine='xlrd')
                    records = df.fillna('').to_dict('records')
                    
                    _logger.info('Excel parsed successfully: %d records found', len(records))
                    if records:
                        _logger.info('Excel headers: %s', list(records[0].keys()))
                    
                    return records
                    
                except ImportError:
                    # Fallback to xlrd only
                    _logger.warning('pandas not available, falling back to xlrd only')
                    return self._parse_excel_xlrd(file_data)
                except Exception as e:
                    _logger.warning('pandas failed, falling back to xlrd: %s', str(e))
                    return self._parse_excel_xlrd(file_data)
            else:
                raise UserError(_('Unsupported Excel format. Please use .xlsx or .xls files.'))

        except UserError:
            # Re-raise UserError as-is
            raise
        except Exception as e:
            _logger.error('Unexpected error parsing Excel: %s', str(e))
            raise UserError(_('Error parsing Excel file: %s') % str(e))
    
    def _parse_excel_xlrd(self, file_data):
        """Parse Excel file using xlrd (fallback method)."""
        try:
            import xlrd
            
            # Create a BytesIO object from the file data
            excel_file = io.BytesIO(file_data)
            
            # Open workbook
            workbook = xlrd.open_workbook(file_contents=excel_file.read())
            sheet = workbook.sheet_by_index(0)  # Use first sheet
            
            # Get headers from first row
            headers = [str(cell.value) for cell in sheet.row(0)]
            _logger.info('Excel headers (xlrd): %s', headers)
            
            # Convert rows to dictionaries
            records = []
            for row_idx in range(1, sheet.nrows):  # Skip header row
                row_data = {}
                for col_idx, header in enumerate(headers):
                    cell_value = sheet.cell(row_idx, col_idx).value
                    # Convert cell value to string, handle empty cells
                    row_data[header] = str(cell_value) if cell_value else ''
                records.append(row_data)
            
            _logger.info('Excel parsed successfully (xlrd): %d records found', len(records))
            return records
            
        except Exception as e:
            _logger.error('Error parsing Excel with xlrd: %s', str(e))
            raise UserError(_('Error parsing Excel file with xlrd: %s') % str(e))
    
    def _validate_records(self, records):
        """Validate records and return list of errors with detailed feedback."""
        errors = []
        warnings = []
        
        # Required fields (updated for Phase 1.1 enhancements)
        required_fields = ['name', 'name_arabic', 'name_english', 'email']
        optional_fields = ['phone', 'birth_date', 'gender', 'nationality', 'native_language', 'english_level', 'has_certificate']
        
        # Track duplicate emails and names
        email_set = set()
        name_set = set()
        
        for i, record in enumerate(records, 1):
            # Check required fields
            for field in required_fields:
                if not record.get(field) or str(record.get(field)).strip() == '':
                    errors.append(f'Row {i}: Missing required field "{field}"')
            
            # Validate email format and uniqueness
            email = record.get('email', '').strip()
            if email:
                if '@' not in email or '.' not in email.split('@')[-1]:
                    errors.append(f'Row {i}: Invalid email format "{email}"')
                elif email.lower() in email_set:
                    errors.append(f'Row {i}: Duplicate email address "{email}"')
                else:
                    email_set.add(email.lower())
            
            # Validate name uniqueness (warning, not error)
            name = record.get('name', '').strip()
            if name:
                if name in name_set:
                    warnings.append(f'Row {i}: Duplicate name "{name}" (may be intentional)')
                else:
                    name_set.add(name)
            
            # Validate birth_date format if provided
            birth_date = record.get('birth_date', '').strip()
            if birth_date:
                try:
                    # Try to parse common date formats
                    datetime.strptime(birth_date, '%Y-%m-%d')
                except ValueError:
                    try:
                        datetime.strptime(birth_date, '%d/%m/%Y')
                    except ValueError:
                        try:
                            datetime.strptime(birth_date, '%m/%d/%Y')
                        except ValueError:
                            errors.append(f'Row {i}: Invalid date format for birth_date "{birth_date}". Use YYYY-MM-DD format.')
            
            # Validate gender if provided
            gender = record.get('gender', '').strip().lower()
            if gender and gender not in ['male', 'female', 'm', 'f']:
                errors.append(f'Row {i}: Invalid gender "{record.get("gender")}". Use "male", "female", "m", or "f".')
            
            # Validate english_level if provided
            english_level = record.get('english_level', '').strip().lower()
            valid_levels = ['beginner', 'elementary', 'intermediate', 'upper_intermediate', 'advanced']
            if english_level and english_level not in valid_levels:
                errors.append(f'Row {i}: Invalid english_level "{record.get("english_level")}". Valid options: {", ".join(valid_levels)}')
            
            # Validate has_certificate if provided
            has_cert = record.get('has_certificate', '').strip().lower()
            if has_cert and has_cert not in ['true', 'false', 'yes', 'no', '1', '0']:
                errors.append(f'Row {i}: Invalid has_certificate value "{record.get("has_certificate")}". Use "true"/"false" or "yes"/"no".')
            
            # Validate certificate consistency
            if has_cert in ['true', 'yes', '1']:
                cert_type = record.get('certificate_type', '').strip()
                cert_date = record.get('certificate_date', '').strip()
                if not cert_type and not cert_date:
                    warnings.append(f'Row {i}: Has certificate is true but no certificate details provided')
                elif cert_date:
                    try:
                        datetime.strptime(cert_date, '%Y-%m-%d')
                    except ValueError:
                        errors.append(f'Row {i}: Invalid certificate_date format "{cert_date}". Use YYYY-MM-DD format.')
            
            # Validate birth date reasonableness (warning)
            if birth_date:
                try:
                    parsed_date = datetime.strptime(birth_date, '%Y-%m-%d')
                    current_year = datetime.now().year
                    if parsed_date.year < 1900 or parsed_date.year > current_year:
                        warnings.append(f'Row {i}: Birth date "{birth_date}" seems unusual')
                except ValueError:
                    pass  # Already handled above
        
        # Store warnings for later reference (Phase 2.3 enhancement)
        if warnings:
            self.validation_warnings = '\n'.join(warnings)
        else:
            self.validation_warnings = False
        
        return errors
    
    def _validate_records_with_details(self, records):
        """Validate records and return detailed error information for failed records management."""
        errors = []
        warnings = []
        failed_records = []
        
        # Required fields (updated for Phase 1.1 enhancements)
        required_fields = ['name', 'name_arabic', 'name_english', 'email']
        optional_fields = ['phone', 'birth_date', 'gender', 'nationality', 'native_language', 'english_level', 'has_certificate']
        
        # Track duplicate emails and names
        email_set = set()
        name_set = set()
        
        for i, record in enumerate(records, 1):
            record_errors = []
            record_warnings = []
            
            # Check required fields
            for field in required_fields:
                if not record.get(field) or str(record.get(field)).strip() == '':
                    error_msg = f'Missing required field "{field}"'
                    errors.append(f'Row {i}: {error_msg}')
                    record_errors.append(error_msg)
            
            # Validate email format and uniqueness
            email = record.get('email', '').strip()
            if email:
                if '@' not in email or '.' not in email.split('@')[-1]:
                    error_msg = f'Invalid email format "{email}"'
                    errors.append(f'Row {i}: {error_msg}')
                    record_errors.append(error_msg)
                elif email.lower() in email_set:
                    error_msg = f'Duplicate email address "{email}"'
                    errors.append(f'Row {i}: {error_msg}')
                    record_errors.append(error_msg)
                else:
                    email_set.add(email.lower())
            
            # Validate name uniqueness (warning, not error)
            name = record.get('name', '').strip()
            if name:
                if name in name_set:
                    warning_msg = f'Duplicate name "{name}" (may be intentional)'
                    warnings.append(f'Row {i}: {warning_msg}')
                    record_warnings.append(warning_msg)
                else:
                    name_set.add(name)
            
            # Validate birth_date format if provided
            birth_date = record.get('birth_date', '').strip()
            if birth_date:
                try:
                    # Try to parse common date formats
                    datetime.strptime(birth_date, '%Y-%m-%d')
                except ValueError:
                    try:
                        datetime.strptime(birth_date, '%d/%m/%Y')
                    except ValueError:
                        try:
                            datetime.strptime(birth_date, '%m/%d/%Y')
                        except ValueError:
                            error_msg = f'Invalid date format for birth_date "{birth_date}". Use YYYY-MM-DD format.'
                            errors.append(f'Row {i}: {error_msg}')
                            record_errors.append(error_msg)
            
            # Validate gender if provided
            gender = record.get('gender', '').strip().lower()
            if gender and gender not in ['male', 'female', 'm', 'f']:
                error_msg = f'Invalid gender "{record.get("gender")}". Use "male", "female", "m", or "f".'
                errors.append(f'Row {i}: {error_msg}')
                record_errors.append(error_msg)
            
            # Validate english_level if provided
            english_level = record.get('english_level', '').strip().lower()
            valid_levels = ['beginner', 'elementary', 'intermediate', 'upper_intermediate', 'advanced']
            if english_level and english_level not in valid_levels:
                error_msg = f'Invalid english_level "{record.get("english_level")}". Valid options: {", ".join(valid_levels)}'
                errors.append(f'Row {i}: {error_msg}')
                record_errors.append(error_msg)
            
            # Validate has_certificate if provided
            has_cert = record.get('has_certificate', '').strip().lower()
            if has_cert and has_cert not in ['true', 'false', 'yes', 'no', '1', '0']:
                error_msg = f'Invalid has_certificate value "{record.get("has_certificate")}". Use "true"/"false" or "yes"/"no".'
                errors.append(f'Row {i}: {error_msg}')
                record_errors.append(error_msg)
            
            # Validate certificate consistency
            if has_cert in ['true', 'yes', '1']:
                cert_type = record.get('certificate_type', '').strip()
                cert_date = record.get('certificate_date', '').strip()
                if not cert_type and not cert_date:
                    warning_msg = f'Has certificate is true but no certificate details provided'
                    warnings.append(f'Row {i}: {warning_msg}')
                    record_warnings.append(warning_msg)
                elif cert_date:
                    try:
                        datetime.strptime(cert_date, '%Y-%m-%d')
                    except ValueError:
                        error_msg = f'Invalid certificate_date format "{cert_date}". Use YYYY-MM-DD format.'
                        errors.append(f'Row {i}: {error_msg}')
                        record_errors.append(error_msg)
            
            # Validate birth date reasonableness (warning)
            if birth_date:
                try:
                    parsed_date = datetime.strptime(birth_date, '%Y-%m-%d')
                    current_year = datetime.now().year
                    if parsed_date.year < 1900 or parsed_date.year > current_year:
                        warning_msg = f'Birth date "{birth_date}" seems unusual'
                        warnings.append(f'Row {i}: {warning_msg}')
                        record_warnings.append(warning_msg)
                except ValueError:
                    pass  # Already handled above
            
            # Store failed record if it has errors
            if record_errors:
                failed_records.append({
                    'row_number': i,
                    'data': record,
                    'errors': record_errors,
                    'warnings': record_warnings,
                    'status': 'failed'
                })
        
        # Store warnings for later reference
        if warnings:
            self.validation_warnings = '\n'.join(warnings)
        else:
            self.validation_warnings = False
        
        # Store failed records data for correction interface
        if failed_records:
            failed_data = {
                'failed_records': failed_records,
                'total_failed': len(failed_records),
                'validation_timestamp': fields.Datetime.now().isoformat(),
                'batch_id': self.id,
                'batch_name': self.name
            }
            self.failed_records_data = json.dumps(failed_data, indent=2)
        else:
            self.failed_records_data = False
        
        return errors
    
    def action_download_template(self):
        """Download a sample Excel template for student data import with validation rules."""
        try:
            # Create sample data with comprehensive examples
            sample_data = [
                {
                    'name': 'Ahmed Hassan',
                    'name_arabic': 'Ahmed Hassan Arabic',
                    'name_english': 'Ahmed Hassan',
                    'email': 'ahmed.hassan@example.com',
                    'phone': '+966501234567',
                    'birth_date': '1995-03-15',
                    'gender': 'male',
                    'nationality': 'Saudi',
                    'native_language': 'Arabic',
                    'english_level': 'intermediate',
                    'has_certificate': 'true',
                    'certificate_type': 'IELTS',
                    'certificate_date': '2023-01-15'
                },
                {
                    'name': 'Fatima Al-Zahra',
                    'name_arabic': 'Fatima Al-Zahra Arabic',
                    'name_english': 'Fatima Al-Zahra',
                    'email': 'fatima.alzahra@example.com',
                    'phone': '+966501234568',
                    'birth_date': '1998-07-22',
                    'gender': 'female',
                    'nationality': 'Saudi',
                    'native_language': 'Arabic',
                    'english_level': 'beginner',
                    'has_certificate': 'false',
                    'certificate_type': '',
                    'certificate_date': ''
                },
                {
                    'name': 'Mohammed Ali',
                    'name_arabic': 'Mohammed Ali Arabic',
                    'name_english': 'Mohammed Ali',
                    'email': 'mohammed.ali@example.com',
                    'phone': '+966501234569',
                    'birth_date': '1993-11-08',
                    'gender': 'male',
                    'nationality': 'Saudi',
                    'native_language': 'Arabic',
                    'english_level': 'advanced',
                    'has_certificate': 'true',
                    'certificate_type': 'TOEFL',
                    'certificate_date': '2022-12-10'
                }
            ]
            
            # Try to create Excel file using pandas if available
            try:
                import pandas as pd
                import io
                
                # Create DataFrame
                df = pd.DataFrame(sample_data)
                
                # Create Excel file in memory with validation rules
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Students', index=False)
                    
                    # Get the workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['Students']
                    
                    # Add validation rules and formatting
                    self._add_excel_validation_rules(workbook, worksheet, df.columns)
                
                # Get the Excel data
                excel_data = output.getvalue()
                output.close()
                
                # Encode as base64
                import base64
                excel_base64 = base64.b64encode(excel_data)
                
                # Create attachment
                attachment = self.env['ir.attachment'].create({
                    'name': 'Student_Import_Template.xlsx',
                    'type': 'binary',
                    'datas': excel_base64,
                    'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'res_model': 'gr.intake.batch',
                    'res_id': self.id,
                })
                
                # Return download action
                return {
                    'type': 'ir.actions.act_url',
                    'url': f'/web/content/{attachment.id}?download=true',
                    'target': 'new',
                }
                
            except ImportError:
                # Fallback: create CSV template
                import csv
                import io
                
                output = io.StringIO()
                if sample_data:
                    writer = csv.DictWriter(output, fieldnames=sample_data[0].keys())
                    writer.writeheader()
                    writer.writerows(sample_data)
                
                csv_data = output.getvalue().encode('utf-8')
                output.close()
                
                # Encode as base64
                import base64
                csv_base64 = base64.b64encode(csv_data)
                
                # Create attachment
                attachment = self.env['ir.attachment'].create({
                    'name': 'Student_Import_Template.csv',
                    'type': 'binary',
                    'datas': csv_base64,
                    'mimetype': 'text/csv',
                    'res_model': 'gr.intake.batch',
                    'res_id': self.id,
                })
                
                # Return download action
                return {
                    'type': 'ir.actions.act_url',
                    'url': f'/web/content/{attachment.id}?download=true',
                    'target': 'new',
                }
                
        except Exception as e:
            _logger.error('Error creating template: %s', str(e))
            raise UserError(_('Error creating template file: %s') % str(e))
    
    def _create_students(self, records):
        """Create student records from validated data with duplicate detection and statistics."""
        created_students = []
        updated_students = []
        skipped_students = []
        errors = []
        
        _logger.info('Starting to create students from %d records', len(records))
        
        for i, record in enumerate(records, 1):
            try:
                # Parse birth_date if provided
                birth_date = None
                if record.get('birth_date'):
                    try:
                        birth_date = datetime.strptime(record.get('birth_date'), '%Y-%m-%d').date()
                    except ValueError:
                        _logger.warning('Invalid birth_date format for student %d: %s', i, record.get('birth_date'))
                
                # Parse certificate_date if provided
                certificate_date = None
                if record.get('certificate_date'):
                    try:
                        certificate_date = datetime.strptime(record.get('certificate_date'), '%Y-%m-%d').date()
                    except ValueError:
                        _logger.warning('Invalid certificate_date format for student %d: %s', i, record.get('certificate_date'))
                
                # Parse has_certificate boolean
                has_certificate = False
                if record.get('has_certificate'):
                    has_cert_str = record.get('has_certificate').lower().strip()
                    has_certificate = has_cert_str in ['true', '1', 'yes', 'y']
                
                # Parse birth_date with multiple format support
                birth_date = None
                if record.get('birth_date'):
                    birth_date_str = record.get('birth_date').strip()
                    try:
                        birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
                    except ValueError:
                        try:
                            birth_date = datetime.strptime(birth_date_str, '%d/%m/%Y').date()
                        except ValueError:
                            try:
                                birth_date = datetime.strptime(birth_date_str, '%m/%d/%Y').date()
                            except ValueError:
                                _logger.warning('Invalid birth_date format for student %d: %s', i, birth_date_str)
                
                # Check for existing student by email (duplicate detection)
                email = record.get('email')
                existing_student = None
                if email:
                    existing_student = self.env['gr.student'].search([
                        ('email', '=', email)
                    ], limit=1)
                
                # Prepare student values
                student_vals = {
                    'name': record.get('name'),
                    'name_arabic': record.get('name_arabic'),
                    'name_english': record.get('name_english'),
                    'email': email,
                    'phone': record.get('phone'),
                    'birth_date': birth_date,
                    'gender': record.get('gender'),
                    'nationality': record.get('nationality'),
                    'native_language': record.get('native_language'),
                    'english_level': record.get('english_level'),
                    'has_certificate': has_certificate,
                    'certificate_type': record.get('certificate_type'),
                    'certificate_date': certificate_date,
                    'intake_batch_id': self.id,
                    'state': 'draft',
                }
                
                if existing_student:
                    # Update existing student with new data
                    _logger.info('Updating existing student %d: %s (ID: %s)', i, student_vals['name'], existing_student.id)
                    
                    # Remove intake_batch_id from update (keep original batch)
                    update_vals = student_vals.copy()
                    del update_vals['intake_batch_id']
                    
                    existing_student.write(update_vals)
                    updated_students.append(existing_student)
                    _logger.info('Student %d updated successfully', i)
                else:
                    # Create new student
                    _logger.info('Creating new student %d: %s with birth_date=%s, has_certificate=%s, english_level=%s', 
                               i, student_vals['name'], birth_date, has_certificate, record.get('english_level'))
                    student = self.env['gr.student'].create(student_vals)
                    created_students.append(student)
                    _logger.info('Student %d created successfully with ID: %s', i, student.id)
                
            except Exception as e:
                error_msg = f'Row {i}: Error creating student "{record.get("name", "Unknown")}": {str(e)}'
                errors.append(error_msg)
                _logger.error('Error creating student %d from record %s: %s', i, record, str(e))
                continue
        
        # Generate import summary
        total_processed = len(created_students) + len(updated_students)
        total_errors = len(errors)
        total_skipped = len(skipped_students)
        
        _logger.info('Student import completed: %d created, %d updated, %d errors, %d skipped out of %d records', 
                    len(created_students), len(updated_students), total_errors, total_skipped, len(records))
        
        # Store import statistics in the batch
        self._store_import_statistics(created_students, updated_students, errors, skipped_students)
        
        return created_students
    
    def _store_import_statistics(self, created_students, updated_students, errors, skipped_students):
        """Store import statistics in the batch record."""
        self.ensure_one()
        
        # Update counts
        self.created_students_count = len(created_students)
        self.updated_students_count = len(updated_students)
        
        # Store errors
        if errors:
            self.import_errors = '\n'.join(errors)
        else:
            self.import_errors = False
        
        # Create summary
        summary_lines = []
        summary_lines.append(f"IMPORT SUMMARY - {fields.Datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        summary_lines.append("=" * 60)
        summary_lines.append(f"Total Records Processed: {len(created_students) + len(updated_students) + len(errors) + len(skipped_students)}")
        summary_lines.append(f"✅ Students Created: {len(created_students)}")
        summary_lines.append(f"🔄 Students Updated: {len(updated_students)}")
        summary_lines.append(f"❌ Errors: {len(errors)}")
        summary_lines.append(f"⏭️ Skipped: {len(skipped_students)}")
        summary_lines.append("")
        
        if created_students:
            summary_lines.append("NEW STUDENTS CREATED:")
            for student in created_students[:10]:  # Show first 10
                summary_lines.append(f"  • {student.name} ({student.email})")
            if len(created_students) > 10:
                summary_lines.append(f"  ... and {len(created_students) - 10} more")
            summary_lines.append("")
        
        if updated_students:
            summary_lines.append("EXISTING STUDENTS UPDATED:")
            for student in updated_students[:10]:  # Show first 10
                summary_lines.append(f"  • {student.name} ({student.email})")
            if len(updated_students) > 10:
                summary_lines.append(f"  ... and {len(updated_students) - 10} more")
            summary_lines.append("")
        
        if errors:
            summary_lines.append("ERRORS ENCOUNTERED:")
            for error in errors[:5]:  # Show first 5 errors
                summary_lines.append(f"  • {error}")
            if len(errors) > 5:
                summary_lines.append(f"  ... and {len(errors) - 5} more errors")
        
        self.import_summary = '\n'.join(summary_lines)
    
    # ===== EXCEL TEMPLATE VALIDATION METHODS =====
    
    def _add_excel_validation_rules(self, workbook, worksheet, columns):
        """Add validation rules and formatting to Excel template."""
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter
            
            # Define validation rules for each field
            validation_rules = {
                'gender': ['male', 'female'],
                'english_level': ['beginner', 'intermediate', 'advanced'],
                'has_certificate': ['true', 'false', 'yes', 'no', '1', '0'],
                'certificate_type': ['IELTS', 'TOEFL', 'Cambridge', 'Other'],
                'nationality': ['Saudi', 'Egyptian', 'Jordanian', 'Lebanese', 'Other'],
                'native_language': ['Arabic', 'English', 'French', 'Other']
            }
            
            # Style for headers
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            # Style for required fields
            required_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            
            # Required fields
            required_fields = ['name', 'name_arabic', 'name_english', 'email']
            
            # Apply formatting to headers and data
            for col_idx, column in enumerate(columns, 1):
                col_letter = get_column_letter(col_idx)
                
                # Header formatting
                header_cell = worksheet[f'{col_letter}1']
                header_cell.fill = header_fill
                header_cell.font = header_font
                header_cell.alignment = Alignment(horizontal='center')
                
                # Required field highlighting
                if column in required_fields:
                    for row in range(2, len(columns) + 4):  # Apply to sample data rows
                        cell = worksheet[f'{col_letter}{row}']
                        cell.fill = required_fill
                
                # Add data validation if rules exist
                if column in validation_rules:
                    validation = DataValidation(
                        type='list',
                        formula1=f'"{",".join(validation_rules[column])}"',
                        allow_blank=True,
                        showErrorMessage=True,
                        errorTitle='Invalid Value',
                        error=f'Please select from: {", ".join(validation_rules[column])}'
                    )
                    worksheet.add_data_validation(validation)
                    validation.add(f'{col_letter}2:{col_letter}1000')  # Apply to data rows
                
                # Auto-adjust column width
                max_length = max(len(str(column)), 15)
                worksheet.column_dimensions[col_letter].width = min(max_length + 2, 30)
            
            # Add instructions sheet
            self._add_instructions_sheet(workbook)
            
        except Exception as e:
            _logger.warning('Could not add Excel validation rules: %s', str(e))
    
    def _add_instructions_sheet(self, workbook):
        """Add instructions sheet to Excel template."""
        try:
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Create instructions sheet
            instructions_sheet = workbook.create_sheet('Instructions', 0)
            
            instructions = [
                ['STUDENT IMPORT TEMPLATE - INSTRUCTIONS'],
                [''],
                ['REQUIRED FIELDS (Highlighted in yellow):'],
                ['• name: Student name in English (required)'],
                ['• name_arabic: Student name in Arabic (required)'],
                ['• name_english: Student name in English - Alternative (required)'],
                ['• email: Valid email address (required)'],
                [''],
                ['OPTIONAL FIELDS:'],
                ['• phone: Contact phone number'],
                ['• birth_date: Date of birth (YYYY-MM-DD format)'],
                ['• gender: male or female'],
                ['• nationality: Student nationality'],
                ['• native_language: Primary language'],
                ['• english_level: beginner, intermediate, or advanced'],
                ['• has_certificate: true/false, yes/no, or 1/0'],
                ['• certificate_type: Type of certificate (if applicable)'],
                ['• certificate_date: Date certificate was obtained'],
                [''],
                ['VALIDATION RULES:'],
                ['• Email addresses must be valid format'],
                ['• Birth dates must be in YYYY-MM-DD format'],
                ['• Gender must be "male" or "female"'],
                ['• English level must be "beginner", "intermediate", or "advanced"'],
                ['• Certificate fields are optional but must be consistent'],
                [''],
                ['IMPORTANT NOTES:'],
                ['• Do not modify column headers'],
                ['• Leave empty cells blank (do not enter "N/A" or similar)'],
                ['• Date format: YYYY-MM-DD (e.g., 1995-03-15)'],
                ['• Boolean values: use "true"/"false", "yes"/"no", or "1"/"0"'],
                ['• Save as .xlsx format for best compatibility'],
            ]
            
            # Add instructions to sheet
            for row_idx, instruction in enumerate(instructions, 1):
                cell = instructions_sheet.cell(row=row_idx, column=1, value=instruction)
                
                if row_idx == 1:  # Title
                    cell.font = Font(size=14, bold=True, color='366092')
                    cell.fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
                elif instruction.startswith('•'):  # Bullet points
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(indent=1)
                elif instruction and not instruction.startswith(' '):  # Headers
                    cell.font = Font(size=11, bold=True)
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            # Auto-adjust column width
            instructions_sheet.column_dimensions['A'].width = 80
            
        except Exception as e:
            _logger.warning('Could not add instructions sheet: %s', str(e))
    
    # ===== LIBRARY CHECK METHODS =====
    
    def _check_excel_libraries(self):
        """Check if required Excel libraries are available."""
        try:
            import pandas as pd
            pandas_available = True
        except ImportError:
            pandas_available = False
            
        try:
            import openpyxl
            openpyxl_available = True
        except ImportError:
            openpyxl_available = False
            
        try:
            import xlrd
            xlrd_available = True
        except ImportError:
            xlrd_available = False
            
        return {
            'pandas': pandas_available,
            'openpyxl': openpyxl_available,
            'xlrd': xlrd_available
        }
    
    def action_check_libraries(self):
        """Check and display library availability."""
        self.ensure_one()
        
        libs = self._check_excel_libraries()
        
        message = "Library Status:\n"
        message += f"• pandas: {'✅ Available' if libs['pandas'] else '❌ Not installed'}\n"
        message += f"• openpyxl: {'✅ Available' if libs['openpyxl'] else '❌ Not installed'}\n"
        message += f"• xlrd: {'✅ Available' if libs['xlrd'] else '❌ Not installed'}\n\n"
        
        if not libs['pandas'] or not libs['openpyxl']:
            message += "⚠️ For .xlsx files, install: pip install pandas openpyxl\n"
        if not libs['xlrd']:
            message += "⚠️ For .xls files, install: pip install xlrd\n"
            
        message += "\n💡 CSV files work without additional libraries."
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Library Check'),
                'message': message,
                'type': 'info',
                'sticky': True,
            }
        }
    
    # ===== STUDENT IMPORT METHODS (Phase 2.4) =====
    
    def action_view_imported_students(self):
        """View students imported from this batch."""
        self.ensure_one()
        
        students = self.env['gr.student'].search([('intake_batch_id', '=', self.id)])
        
        return {
            'name': _('Imported Students - %s') % self.name,
            'type': 'ir.actions.act_window',
            'res_model': 'gr.student',
            'view_mode': 'list,form',
            'domain': [('id', 'in', students.ids)],
            'context': {
                'default_intake_batch_id': self.id,
                'search_default_intake_batch_id': self.id,
            },
            'target': 'current',
        }
    
    def action_view_created_students(self):
        """View only newly created students from this batch."""
        self.ensure_one()
        
        # Get students created in this batch (not updated ones)
        students = self.env['gr.student'].search([
            ('intake_batch_id', '=', self.id),
            ('create_date', '>=', self.processing_date or self.upload_date)
        ])
        
        return {
            'name': _('New Students Created - %s') % self.name,
            'type': 'ir.actions.act_window',
            'res_model': 'gr.student',
            'view_mode': 'list,form',
            'domain': [('id', 'in', students.ids)],
            'context': {
                'default_intake_batch_id': self.id,
                'search_default_intake_batch_id': self.id,
            },
            'target': 'current',
        }
    
    def action_process_large_batch(self):
        """Process large datasets in batches to avoid memory issues."""
        self.ensure_one()
        
        if self.state != 'validated':
            raise UserError(_('Please validate the file first.'))
        
        # Check if this is a large batch (>1000 records)
        if self.total_records <= 1000:
            return self.action_process_file()
        
        # For large batches, show confirmation dialog
        return {
            'name': _('Process Large Batch'),
            'type': 'ir.actions.act_window',
            'res_model': 'gr.intake.batch.large.process.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_intake_batch_id': self.id,
                'default_total_records': self.total_records,
            }
        }
    
    # ===== NOTIFICATION METHODS (Phase 3.1.3) =====
    
    def _get_notification_recipients(self):
        """Get list of users who should receive notifications for this batch."""
        recipients = []
        
        # Add batch creator
        if self.create_uid:
            recipients.append(self.create_uid)
        
        # Add managers (users with grants_training_suite_v19.group_manager)
        try:
            manager_group = self.env.ref('grants_training_suite_v19.group_manager')
            managers = self.env['res.users'].search([('groups_id', 'in', manager_group.id)])
            recipients.extend(managers)
        except Exception:
            pass
        
        # Add agents if this batch has assigned agents
        if hasattr(self, 'agent_ids') and self.agent_ids:
            recipients.extend(self.agent_ids)
        
        # Remove duplicates and inactive users
        unique_recipients = []
        seen_users = set()
        for user in recipients:
            if user.active and user.id not in seen_users:
                unique_recipients.append(user)
                seen_users.add(user.id)
        
        return unique_recipients
    
    def _send_batch_notification(self, notification_type, message, details=None):
        """Send notification for batch status change."""
        self.ensure_one()
        
        try:
            recipients = self._get_notification_recipients()
            if not recipients:
                _logger.warning('No recipients found for batch notification: %s', self.name)
                return
            
            # Prepare notification data
            notification_data = {
                'notification_type': notification_type,
                'message': message,
                'details': details or {},
                'batch_name': self.name,
                'batch_id': self.id,
                'timestamp': fields.Datetime.now(),
                'recipients': [{'id': r.id, 'name': r.name, 'email': r.email} for r in recipients]
            }
            
            # Send email notifications
            if self.email_notification_enabled:
                self._send_email_notifications(recipients, notification_data)
            
            # Send in-app notifications
            if self.in_app_notification_enabled:
                self._send_in_app_notifications(recipients, notification_data)
            
            # Update batch notification fields
            self.write({
                'notification_sent': True,
                'notification_type': notification_type,
                'notification_message': message,
                'notification_recipients': ', '.join([r.name for r in recipients]),
                'notification_date': fields.Datetime.now(),
            })
            
            _logger.info('Sent %s notification for batch %s to %d recipients', 
                        notification_type, self.name, len(recipients))
            
        except Exception as e:
            _logger.error('Error sending notification for batch %s: %s', self.name, str(e))
    
    def _send_email_notifications(self, recipients, notification_data):
        """Send email notifications to recipients."""
        try:
            # Get email template based on notification type
            template_name = self._get_email_template_name(notification_data['notification_type'])
            
            try:
                template = self.env.ref(template_name)
            except Exception:
                # Fallback to generic template
                template = self.env.ref('grants_training_suite_v19.email_template_batch_notification_generic')
            
            # Send email to each recipient
            for recipient in recipients:
                if recipient.email:
                    template.send_mail(self.id, force_send=True, email_values={
                        'email_to': recipient.email,
                        'recipient_name': recipient.name,
                        'notification_type': notification_data['notification_type'],
                        'message': notification_data['message'],
                        'batch_name': notification_data['batch_name'],
                        'details': notification_data['details'],
                    })
                    _logger.info('Sent email notification to %s (%s)', recipient.name, recipient.email)
            
        except Exception as e:
            _logger.error('Error sending email notifications: %s', str(e))
    
    def _send_in_app_notifications(self, recipients, notification_data):
        """Send in-app notifications to recipients."""
        try:
            notification_type_mapping = {
                'success': 'success',
                'error': 'danger',
                'warning': 'warning',
                'info': 'info'
            }
            
            notification_type = notification_type_mapping.get(notification_data['notification_type'], 'info')
            
            for recipient in recipients:
                self.env['mail.message'].create({
                    'model': 'gr.intake.batch',
                    'res_id': self.id,
                    'message_type': 'notification',
                    'subtype_id': self.env.ref('mail.mt_note').id,
                    'body': f"<p><strong>Batch {notification_data['batch_name']}</strong></p><p>{notification_data['message']}</p>",
                    'subject': f"Batch Notification: {notification_data['batch_name']}",
                    'author_id': self.env.user.id,
                    'partner_ids': [(4, recipient.partner_id.id)],
                    'notification_ids': [(0, 0, {
                        'res_partner_id': recipient.partner_id.id,
                        'notification_type': 'inbox',
                        'notification_status': 'ready',
                        'is_read': False,
                    })]
                })
            
        except Exception as e:
            _logger.error('Error sending in-app notifications: %s', str(e))
    
    def _get_email_template_name(self, notification_type):
        """Get email template name based on notification type."""
        template_mapping = {
            'success': 'grants_training_suite_v19.email_template_batch_success',
            'error': 'grants_training_suite_v19.email_template_batch_error',
            'warning': 'grants_training_suite_v19.email_template_batch_warning',
            'info': 'grants_training_suite_v19.email_template_batch_info',
        }
        return template_mapping.get(notification_type, 'grants_training_suite_v19.email_template_batch_notification_generic')
    
    def action_send_test_notification(self):
        """Send a test notification for this batch."""
        self.ensure_one()
        
        test_message = f"Test notification for batch '{self.name}'. This is a test message to verify notification functionality."
        
        self._send_batch_notification('info', test_message, {
            'test_mode': True,
            'batch_status': self.state,
            'total_records': self.total_records or 0,
        })
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Test Notification Sent'),
                'message': _('Test notification has been sent to all configured recipients.'),
                'type': 'success',
            }
        }
    
    def action_resend_notification(self):
        """Resend the last notification for this batch."""
        self.ensure_one()
        
        if not self.notification_sent:
            raise UserError(_('No notification has been sent for this batch yet.'))
        
        # Resend with the same message
        self._send_batch_notification(
            self.notification_type,
            self.notification_message,
            {'resend': True}
        )
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Notification Resent'),
                'message': _('Notification has been resent to all configured recipients.'),
                'type': 'success',
            }
        }

    # ===== SESSION AUTOMATION METHODS (Phase 3.2) =====
    
    def action_create_sessions_for_batch(self):
        """Create course sessions for all students in this batch."""
        self.ensure_one()
        
        if self.state != 'processed':
            raise UserError(_('Batch must be processed before creating sessions.'))
        
        if not self.session_creation_enabled:
            raise UserError(_('Session creation is disabled for this batch.'))
        
        try:
            # Get all students from this batch
            students = self.env['gr.student'].search([
                ('intake_batch_id', '=', self.id)
            ])
            
            if not students:
                raise UserError(_('No students found in this batch.'))
            
            created_sessions = []
            errors = []
            session_creation_start = fields.Datetime.now()
            
            for student in students:
                try:
                    session = self._create_session_for_student(student)
                    if session:
                        created_sessions.append(session)
                except Exception as e:
                    error_msg = f'Error creating session for student {student.name}: {str(e)}'
                    errors.append(error_msg)
                    _logger.error(error_msg)
            
            # Update batch with session creation results
            self.write({
                'sessions_created_count': len(created_sessions),
                'sessions_scheduled_count': len([s for s in created_sessions if s.state == 'scheduled']),
                'session_creation_date': session_creation_start,
                'session_creation_errors': '\n'.join(errors) if errors else False,
                'session_creation_summary': self._generate_session_creation_summary(created_sessions, errors)
            })
            
            # Send notification about session creation
            if created_sessions:
                notification_type = 'warning' if errors else 'success'
                message = f"Created {len(created_sessions)} sessions for batch '{self.name}'."
                if errors:
                    message += f" {len(errors)} errors encountered."
                
                self._send_batch_notification(notification_type, message, {
                    'sessions_created': len(created_sessions),
                    'sessions_scheduled': len([s for s in created_sessions if s.state == 'scheduled']),
                    'errors': len(errors),
                    'students_processed': len(students)
                })
            
            _logger.info('Session creation completed for batch %s: %d sessions created, %d errors', 
                        self.name, len(created_sessions), len(errors))
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Sessions Created'),
                    'message': _('Successfully created %d sessions. %d errors encountered.') % 
                              (len(created_sessions), len(errors)),
                    'type': 'warning' if errors else 'success',
                    'sticky': True,
                }
            }
            
        except Exception as e:
            _logger.error('Error creating sessions for batch %s: %s', self.name, str(e))
            raise UserError(_('Error creating sessions: %s') % str(e))
    
    def _create_session_for_student(self, student):
        """Create a course session for a specific student."""
        try:
            # Determine session parameters
            session_date = self._calculate_session_date(student)
            session_duration = self.default_session_duration
            session_type = self.default_session_type
            
            # Use session template if available
            session_vals = {
                'name': f"Session - {student.name} - {session_date.strftime('%Y-%m-%d %H:%M')}",
                'student_id': student.id,
                'session_date': session_date,
                'session_duration': session_duration,
                'session_type': session_type,
                'state': 'scheduled',
                'session_topic': self._get_default_session_topic(student),
                'session_objectives': self._get_default_session_objectives(student),
            }
            
            # Apply template if available
            if self.session_template_id:
                template_vals = self._apply_session_template(session_vals)
                session_vals.update(template_vals)
            
            # Create the session
            session = self.env['gr.course.session'].create(session_vals)
            
            _logger.info('Created session %s for student %s', session.name, student.name)
            return session
            
        except Exception as e:
            _logger.error('Error creating session for student %s: %s', student.name, str(e))
            raise e
    
    def _calculate_session_date(self, student):
        """Calculate appropriate session date for student."""
        from datetime import datetime, timedelta
        
        # Start from next week to allow time for preparation
        base_date = fields.Datetime.now() + timedelta(days=7)
        
        # Use student's preferred schedule if available
        # For now, default to weekday mornings (9 AM)
        session_date = base_date.replace(hour=9, minute=0, second=0, microsecond=0)
        
        # Adjust to next weekday if it's weekend
        if session_date.weekday() >= 5:  # Saturday = 5, Sunday = 6
            days_to_add = 7 - session_date.weekday()
            session_date += timedelta(days=days_to_add)
        
        return session_date
    
    def _get_default_session_topic(self, student):
        """Get default session topic based on student's course preference."""
        if student.preferred_course_integration_id:
            return f"Introduction to {student.preferred_course_integration_id.name}"
        return "Initial Assessment and Goal Setting"
    
    def _get_default_session_objectives(self, student):
        """Get default session objectives based on student's profile."""
        objectives = [
            "Assess student's current skill level",
            "Review learning goals and expectations",
            "Establish communication preferences",
            "Create personalized learning plan"
        ]
        
        if student.preferred_course_integration_id:
            objectives.append(f"Introduce {student.preferred_course_integration_id.name} curriculum")
        
        return '\n'.join(objectives)
    
    def _apply_session_template(self, session_vals):
        """Apply session template to session values."""
        template_vals = {}
        
        if self.session_template_id:
            # Apply template-specific values
            template_vals.update({
                'session_topic': self.session_template_id.default_topic or session_vals.get('session_topic'),
                'session_objectives': self.session_template_id.default_objectives or session_vals.get('session_objectives'),
                'session_duration': self.session_template_id.default_duration or session_vals.get('session_duration'),
                'session_type': self.session_template_id.default_type or session_vals.get('session_type'),
                'location': self.session_template_id.default_location,
                'meeting_link': self.session_template_id.default_meeting_link,
            })
        
        return template_vals
    
    def _generate_session_creation_summary(self, created_sessions, errors):
        """Generate summary of session creation process."""
        summary_lines = [
            f"Session Creation Summary for Batch: {self.name}",
            f"Date: {fields.Datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"Total Sessions Created: {len(created_sessions)}",
            f"Total Errors: {len(errors)}",
            ""
        ]
        
        if created_sessions:
            summary_lines.append("Created Sessions:")
            for session in created_sessions:
                summary_lines.append(f"  - {session.name} ({session.student_id.name})")
            summary_lines.append("")
        
        if errors:
            summary_lines.append("Errors:")
            for error in errors:
                summary_lines.append(f"  - {error}")
        
        return '\n'.join(summary_lines)
    
    def action_view_created_sessions(self):
        """View sessions created from this batch."""
        self.ensure_one()
        
        sessions = self.env['gr.course.session'].search([
            ('student_id.intake_batch_id', '=', self.id)
        ])
        
        return {
            'name': _('Sessions Created from Batch %s') % self.name,
            'type': 'ir.actions.act_window',
            'res_model': 'gr.course.session',
            'view_mode': 'list,form',
            'domain': [('student_id.intake_batch_id', '=', self.id)],
            'context': {
                'default_student_id': False,
                'search_default_scheduled': 1,
            }
        }

    # ===== FAILED RECORDS MANAGEMENT METHODS (Phase 3.1.2) =====
    
    def action_view_failed_records(self):
        """View failed records for this batch."""
        self.ensure_one()
        
        if not self.has_failed_records:
            raise UserError(_('No failed records found for this batch.'))
        
        return {
            'name': _('Failed Records - %s') % self.name,
            'type': 'ir.actions.act_window',
            'res_model': 'gr.intake.batch.correction.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_intake_batch_id': self.id,
                'default_batch_name': self.name,
            }
        }
    
    def action_open_correction_wizard(self):
        """Open correction wizard for failed records."""
        self.ensure_one()
        
        if not self.has_failed_records:
            raise UserError(_('No failed records found for this batch.'))
        
        # Create new correction wizard
        wizard = self.env['gr.intake.batch.correction.wizard'].create({
            'intake_batch_id': self.id,
            'batch_name': self.name,
        })
        
        return {
            'name': _('Correct Failed Records - %s') % self.name,
            'type': 'ir.actions.act_window',
            'res_model': 'gr.intake.batch.correction.wizard',
            'view_mode': 'form',
            'res_id': wizard.id,
            'target': 'new',
            'context': {
                'default_intake_batch_id': self.id,
                'default_batch_name': self.name,
            }
        }
    
    def action_reprocess_failed_records(self):
        """Re-process only the failed records after corrections."""
        self.ensure_one()
        
        if not self.has_failed_records:
            raise UserError(_('No failed records found for this batch.'))
        
        # For now, we'll show a message that this feature requires opening the correction wizard first
        # In a full implementation, we could store the corrected data in a regular model or use a different approach
        raise UserError(_('Please open the correction wizard first to correct the failed records, then process them from there.'))
    
    def _create_students_from_corrected_data(self, corrected_records):
        """Create students from corrected record data."""
        created_students = []
        
        for record_data in corrected_records:
            try:
                # Extract student data from corrected record
                student_vals = record_data.get('data', {})
                
                # Ensure required fields are present
                if not student_vals.get('name_arabic') or not student_vals.get('name_english'):
                    student_vals['name_arabic'] = student_vals.get('name', '')
                    student_vals['name_english'] = student_vals.get('name', '')
                
                # Set intake batch
                student_vals['intake_batch_id'] = self.id
                student_vals['intake_date'] = fields.Datetime.now()
                
                # Create student
                student = self.env['gr.student'].create(student_vals)
                created_students.append(student)
                
            except Exception as e:
                _logger.error('Error creating student from corrected record: %s', str(e))
                continue
        
        return created_students
    
    def action_validate_with_failed_tracking(self):
        """Validate records with detailed failed records tracking."""
        self.ensure_one()
        
        if self.state not in ['uploaded', 'mapping']:
            raise UserError(_('Please upload a file first.'))
        
        try:
            # Set validation progress to in_progress
            self.validation_progress = 'in_progress'
            
            # Parse file
            records = self._parse_file()
            self.total_records = len(records)
            
            # Validate with detailed tracking
            errors = self._validate_records_with_details(records)
            
            if errors:
                self.validation_errors = '\n'.join(errors)
                self.state = 'error'
                self.validation_progress = 'failed'
                self.error_records = len(errors)
            else:
                self.validation_errors = False
                self.state = 'validated'
                self.validation_progress = 'completed'
                self.error_records = 0
            
            self.validation_date = fields.Datetime.now()
            
            # Log the validation
            _logger.info('File validated with failed tracking for batch %s: %d records, %d errors', 
                        self.name, self.total_records, self.error_records)
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('File Validated'),
                    'message': _('File validation completed. %d records found, %d errors.') % 
                              (self.total_records, self.error_records),
                    'type': 'success' if not errors else 'warning',
                }
            }
            
        except Exception as e:
            _logger.error('Error validating batch %s: %s', self.name, str(e))
            self.state = 'error'
            self.validation_progress = 'failed'
            self.validation_errors = str(e)
            raise UserError(_('Error validating file: %s') % str(e))

    # ===== REAL-TIME VALIDATION METHODS (Phase 2.3) =====
    
    def action_validate_preview(self):
        """Provide real-time validation preview without changing state."""
        self.ensure_one()
        
        if not self.file_data:
            raise UserError(_('Please upload a file first.'))
        
        try:
            # Parse file
            if self.file_type == 'csv':
                records = self._parse_csv(self.file_data)
            else:
                records = self._parse_excel(self.file_data)
            
            if not records:
                raise UserError(_('No data found in the uploaded file.'))
            
            # Validate records (this will populate warnings)
            errors = self._validate_records(records)
            
            # Prepare feedback message
            total_records = len(records)
            error_count = len(errors)
            warning_count = len(self.validation_warnings.split('\n')) if self.validation_warnings else 0
            
            message = f"Validation Preview:\n"
            message += f"• Total records: {total_records}\n"
            message += f"• Errors: {error_count}\n"
            message += f"• Warnings: {warning_count}\n\n"
            
            if error_count > 0:
                message += f"❌ {error_count} errors found. Please fix these before proceeding:\n"
                for error in errors[:5]:  # Show first 5 errors
                    message += f"  - {error}\n"
                if error_count > 5:
                    message += f"  ... and {error_count - 5} more errors\n"
            else:
                message += "✅ No validation errors found!\n"
            
            if warning_count > 0:
                message += f"\n⚠️ {warning_count} warnings found:\n"
                warnings_list = self.validation_warnings.split('\n')[:3]  # Show first 3 warnings
                for warning in warnings_list:
                    message += f"  - {warning}\n"
                if warning_count > 3:
                    message += f"  ... and {warning_count - 3} more warnings\n"
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Validation Preview'),
                    'message': message,
                    'type': 'warning' if error_count > 0 else 'success',
                    'sticky': True,
                }
            }
            
        except Exception as e:
            _logger.error('Error in validation preview: %s', str(e))
            raise UserError(_('Error validating file: %s') % str(e))
    
    def action_show_validation_details(self):
        """Show detailed validation results."""
        self.ensure_one()
        
        if not self.validation_errors and not self.validation_warnings:
            raise UserError(_('No validation results available. Please validate the file first.'))
        
        # Create a detailed report
        report = "VALIDATION REPORT\n"
        report += "=" * 50 + "\n\n"
        
        if self.validation_errors:
            report += "ERRORS:\n"
            report += "-" * 20 + "\n"
            report += self.validation_errors + "\n\n"
        
        if self.validation_warnings:
            report += "WARNINGS:\n"
            report += "-" * 20 + "\n"
            report += self.validation_warnings + "\n\n"
        
        return {
            'type': 'ir.actions.act_window',
            'name': _('Validation Details'),
            'res_model': 'gr.intake.batch',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
            'context': {
                'show_validation_report': True,
                'validation_report': report
            }
        }
    
    # ===== COLUMN MAPPING METHODS (Phase 2.2) =====
    
    def _get_student_field_mapping(self):
        """Get the mapping of available student fields for column mapping."""
        return {
            'name': {
                'label': 'Student Name (English)',
                'required': True,
                'help': 'Full name in English'
            },
            'name_arabic': {
                'label': 'Student Name (Arabic)',
                'required': True,
                'help': 'Full name in Arabic'
            },
            'name_english': {
                'label': 'Student Name (English) - Alternative',
                'required': True,
                'help': 'Alternative English name field'
            },
            'email': {
                'label': 'Email Address',
                'required': True,
                'help': 'Valid email address'
            },
            'phone': {
                'label': 'Phone Number',
                'required': False,
                'help': 'Contact phone number'
            },
            'birth_date': {
                'label': 'Birth Date',
                'required': False,
                'help': 'Date of birth (YYYY-MM-DD format)'
            },
            'gender': {
                'label': 'Gender',
                'required': False,
                'help': 'Male or Female'
            },
            'nationality': {
                'label': 'Nationality',
                'required': False,
                'help': 'Student nationality'
            },
            'native_language': {
                'label': 'Native Language',
                'required': False,
                'help': 'Primary language'
            },
            'english_level': {
                'label': 'English Level',
                'required': False,
                'help': 'Beginner, Intermediate, Advanced'
            },
            'has_certificate': {
                'label': 'Has Certificate',
                'required': False,
                'help': 'True/False or Yes/No'
            },
            'certificate_type': {
                'label': 'Certificate Type',
                'required': False,
                'help': 'Type of certificate if applicable'
            },
            'certificate_date': {
                'label': 'Certificate Date',
                'required': False,
                'help': 'Date certificate was obtained'
            }
        }
    
    def action_open_column_mapping(self):
        """Open column mapping wizard."""
        self.ensure_one()
        
        if not self.file_data:
            raise UserError(_('Please upload a file first.'))
        
        # Set mapping progress to in_progress
        self.mapping_progress = 'in_progress'
        
        # Parse file to get available columns
        try:
            if self.file_type == 'csv':
                records = self._parse_csv(self.file_data)
            else:
                records = self._parse_excel(self.file_data)
            
            if not records:
                self.mapping_progress = 'failed'
                raise UserError(_('No data found in the uploaded file.'))
            
            # Get available columns from the first record
            available_columns = list(records[0].keys()) if records else []
            
            # Get sample data (first 3 records)
            preview_data = records[:3] if len(records) >= 3 else records
            
            # Auto-detect mapping based on column names
            auto_mapping = self._auto_detect_column_mapping(available_columns)
            
            # Store data for the wizard
            self.write({
                'available_columns': json.dumps(available_columns),
                'mapping_preview_data': json.dumps(preview_data),
                'column_mapping': json.dumps(auto_mapping),
                'state': 'mapping',
                'mapping_progress': 'completed'
            })
            
            # Return wizard action
            return {
                'type': 'ir.actions.act_window',
                'name': _('Column Mapping'),
                'res_model': 'gr.intake.batch.mapping.wizard',
                'view_mode': 'form',
                'target': 'new',
                'context': {
                    'default_intake_batch_id': self.id,
                    'default_available_columns': json.dumps(available_columns),
                    'default_preview_data': json.dumps(preview_data),
                    'default_column_mapping': json.dumps(auto_mapping),
                }
            }
            
        except Exception as e:
            _logger.error('Error opening column mapping: %s', str(e))
            self.mapping_progress = 'failed'
            raise UserError(_('Error parsing file for column mapping: %s') % str(e))
    
    def _auto_detect_column_mapping(self, available_columns):
        """Auto-detect column mapping based on column names."""
        mapping = {}
        student_fields = self._get_student_field_mapping()
        
        # Common column name patterns
        patterns = {
            'name': ['name', 'full_name', 'student_name', 'fullname', 'english_name'],
            'name_arabic': ['arabic_name', 'arabic', 'name_arabic', 'arabic_full_name'],
            'name_english': ['english_name', 'english', 'name_english', 'english_full_name'],
            'email': ['email', 'email_address', 'e_mail', 'mail'],
            'phone': ['phone', 'phone_number', 'mobile', 'telephone', 'contact_number'],
            'birth_date': ['birth_date', 'birthdate', 'date_of_birth', 'dob', 'birth'],
            'gender': ['gender', 'sex'],
            'nationality': ['nationality', 'country', 'citizenship'],
            'native_language': ['native_language', 'language', 'mother_tongue'],
            'english_level': ['english_level', 'english', 'level', 'proficiency'],
            'has_certificate': ['has_certificate', 'certificate', 'cert'],
            'certificate_type': ['certificate_type', 'cert_type', 'certification'],
            'certificate_date': ['certificate_date', 'cert_date', 'certification_date']
        }
        
        # Try to match columns
        for field, patterns_list in patterns.items():
            for col in available_columns:
                col_lower = col.lower().strip().replace(' ', '_').replace('-', '_')
                if any(pattern in col_lower for pattern in patterns_list):
                    mapping[field] = col
                    break
        
        return mapping
    
    def action_save_column_mapping(self, mapping_data):
        """Save column mapping and proceed to validation."""
        self.ensure_one()
        
        try:
            # Validate mapping data
            mapping = json.loads(mapping_data) if isinstance(mapping_data, str) else mapping_data
            
            # Check required fields are mapped
            student_fields = self._get_student_field_mapping()
            required_fields = [field for field, config in student_fields.items() if config['required']]
            missing_fields = [field for field in required_fields if field not in mapping or not mapping[field]]
            
            if missing_fields:
                missing_labels = [student_fields[field]['label'] for field in missing_fields]
                raise UserError(_('Please map the following required fields: %s') % ', '.join(missing_labels))
            
            # Save mapping
            self.column_mapping = json.dumps(mapping)
            
            # Process file with mapping
            self.action_process_with_mapping()
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Success'),
                    'message': _('Column mapping saved successfully. File validation in progress...'),
                    'type': 'success',
                    'sticky': False,
                }
            }
            
        except Exception as e:
            _logger.error('Error saving column mapping: %s', str(e))
            raise UserError(_('Error saving column mapping: %s') % str(e))
    
    def action_process_with_mapping(self):
        """Process file using the column mapping."""
        self.ensure_one()
        
        if not self.column_mapping:
            raise UserError(_('No column mapping found. Please configure column mapping first.'))
        
        try:
            mapping = json.loads(self.column_mapping)
            
            # Parse file
            if self.file_type == 'csv':
                raw_records = self._parse_csv(self.file_data)
            else:
                raw_records = self._parse_excel(self.file_data)
            
            if not raw_records:
                raise UserError(_('No data found in the uploaded file.'))
            
            # Apply mapping to transform records
            mapped_records = []
            for record in raw_records:
                mapped_record = {}
                for field, column in mapping.items():
                    if column and column in record:
                        mapped_record[field] = record[column]
                mapped_records.append(mapped_record)
            
            # Update total records count
            self.total_records = len(mapped_records)
            
            # Validate mapped records
            validation_errors = self._validate_records(mapped_records)
            
            if validation_errors:
                self.state = 'error'
                self.validation_errors = '\n'.join(validation_errors)
                self.error_records = len(validation_errors)
                self.validation_date = fields.Datetime.now()
                
                raise UserError(_('Validation failed. Please check the error details and fix the data.'))
            else:
                self.state = 'validated'
                self.validation_errors = False
                self.error_records = 0
                self.validation_date = fields.Datetime.now()
                
                # Store mapped records for processing
                self.mapping_preview_data = json.dumps(mapped_records[:10])  # Store first 10 for reference
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Success'),
                    'message': _('File validation completed successfully. %d records ready for processing.') % len(mapped_records),
                    'type': 'success',
                    'sticky': False,
                }
            }
            
        except Exception as e:
            _logger.error('Error processing file with mapping: %s', str(e))
            self.state = 'error'
            self.validation_errors = str(e)
            raise UserError(_('Error processing file: %s') % str(e))
    
    def action_reset(self):
        """Reset batch to draft state."""
        self.ensure_one()
        self.state = 'draft'
        self.total_records = 0
        self.processed_records = 0
        self.error_records = 0
        self.validation_errors = False
        self.upload_date = False
        self.validation_date = False
        self.processing_date = False
        # Reset column mapping fields (Phase 2.2)
        self.column_mapping = False
        self.available_columns = False
        self.mapping_preview_data = False
        
        # Reset import statistics fields (Phase 2.4)
        self.created_students_count = 0
        self.updated_students_count = 0
        self.import_errors = False
        self.import_summary = False
        
        # Reset progress tracking fields (Phase 3.1)
        self.upload_progress = 'pending'
        self.mapping_progress = 'pending'
        self.validation_progress = 'pending'
        self.processing_progress = 'pending'
        
        # Reset failed records fields (Phase 3.1.2)
        self.failed_records_data = False
        
        # Reset notification fields (Phase 3.1.3)
        self.notification_sent = False
        self.notification_type = 'none'
        self.notification_message = False
        self.notification_recipients = False
        self.notification_date = False
        
        # Reset session automation fields (Phase 3.2)
        self.sessions_created_count = 0
        self.sessions_scheduled_count = 0
        self.session_creation_date = False
        self.session_creation_errors = False
        self.session_creation_summary = False
